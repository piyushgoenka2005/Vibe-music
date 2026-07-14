import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "Vibe Musice 07-07-2026.xlsx"
PRODUCTS_PATH = ROOT / "products.json"
CATALOG_PATH = ROOT / "src" / "data" / "catalog" / "products.json"


def norm(s: str) -> str:
    s = s.upper()
    s = s.replace("″", '"').replace("”", '"').replace("“", '"')
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def tokens(s: str) -> set[str]:
    return {t for t in norm(s).split() if t and t not in {"ADEON", "AVUS", "THE", "AND", "WITH"}}


def round_price(value: float) -> int:
    return int(round(float(value)))


def score_match(excel_name: str, product: dict) -> int:
    e = norm(excel_name)
    candidates = [
        product.get("name") or "",
        product.get("slug") or "",
        product.get("sku") or "",
        str((product.get("specifications") or {}).get("Model") or ""),
        str((product.get("specifications") or {}).get("Product Type") or ""),
    ]
    best = 0
    e_tokens = tokens(excel_name)
    # Prefer model-like compact forms: ADEON ADM - 01 -> ADM 01 / ADM01
    compact = e.replace(" ", "")
    for c in candidates:
        n = norm(c)
        cpt = n.replace(" ", "")
        if not n:
            continue
        if e in n or n in e:
            best = max(best, 100)
        if compact and (compact in cpt or cpt in compact):
            best = max(best, 95)
        # strip brand from excel for model match
        model_only = re.sub(r"^(ADEON|AVUS|HERTZ|ROLAND|ZOOM)\s+", "", e)
        model_compact = model_only.replace(" ", "")
        if model_compact and len(model_compact) >= 3 and model_compact in cpt:
            best = max(best, 90)
        overlap = len(e_tokens & tokens(c))
        if overlap:
            best = max(best, overlap * 12)
    return best


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    excel_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, pcs, mrp, vibe_price, discount = row[:5]
        if not name:
            continue
        excel_rows.append(
            {
                "name": str(name).strip(),
                "pcs": pcs,
                "mrp": float(mrp or 0),
                "price": float(vibe_price or 0),
                "discount": float(discount or 0),
            }
        )

    products = json.loads(PRODUCTS_PATH.read_text(encoding="utf-8"))
    used_ids: set[str] = set()
    matches = []
    unmatched = []

    for er in excel_rows:
        ranked = sorted(
            ((score_match(er["name"], p), p) for p in products if p["id"] not in used_ids),
            key=lambda x: x[0],
            reverse=True,
        )
        if not ranked or ranked[0][0] < 70:
            unmatched.append(er["name"])
            print(f"NO MATCH  {er['name']!r}  top={ranked[0][0] if ranked else 0} {ranked[0][1].get('slug') if ranked else ''}")
            continue
        score, product = ranked[0]
        used_ids.add(product["id"])
        matches.append((er, product, score))
        print(
            f"MATCH[{score}] {er['name']!r} -> {product.get('sku')} / {product.get('slug')} "
            f"(old price={product.get('price')} -> {round_price(er['price'])})"
        )

    print(f"\nmatched={len(matches)} unmatched={len(unmatched)} excel={len(excel_rows)} products={len(products)}")

    now = "2026-07-14T06:30:00.000Z"
    by_id = {p["id"]: p for p in products}

    for er, product, _score in matches:
        sale = round_price(er["price"])
        mrp = round_price(er["mrp"]) if er["mrp"] else sale
        if mrp < sale:
            mrp = sale
        discount_pct = 0
        if mrp > sale > 0:
            discount_pct = int(round((1 - sale / mrp) * 100))

        product["price"] = sale
        product["originalPrice"] = mrp
        product["discountPercentage"] = discount_pct
        product["status"] = "active"
        product["availability"] = "in-stock"
        # Keep reasonable buyable stock; use sheet Pcs when available
        pcs = int(er["pcs"] or 0)
        product["stock"] = max(pcs, 1) if pcs > 0 else max(int(product.get("stock") or 0), 10)
        product["reservedStock"] = int(product.get("reservedStock") or 0)
        if product["stock"] <= product["reservedStock"]:
            product["reservedStock"] = 0
        product["updatedAt"] = now
        detail = product.get("detail") or {}
        detail["salePrice"] = sale
        detail["msrp"] = mrp
        product["detail"] = detail

        # Sync default variant prices if present
        variants = product.get("variants") or detail.get("variants") or []
        for variant in variants:
            if isinstance(variant, dict):
                if "price" in variant:
                    variant["price"] = sale
                if "compareAtPrice" in variant or "originalPrice" in variant:
                    variant["compareAtPrice"] = mrp
                    variant["originalPrice"] = mrp
                if "availability" in variant:
                    variant["availability"] = "in-stock"
                if "stock" in variant:
                    variant["stock"] = product["stock"]

    PRODUCTS_PATH.write_text(json.dumps(products, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    # Keep src/data/catalog/products.json in sync for IDs present there
    if CATALOG_PATH.exists():
        catalog = json.loads(CATALOG_PATH.read_text(encoding="utf-8"))
        updated = 0
        for i, cp in enumerate(catalog):
            src = by_id.get(cp["id"])
            if not src:
                # try slug
                src = next((p for p in products if p.get("slug") == cp.get("slug")), None)
            if not src:
                continue
            if src["id"] in used_ids or any(m[1]["slug"] == src.get("slug") for m in matches):
                catalog[i] = src
                updated += 1
        # Also ensure matched products exist in catalog: upsert by id
        catalog_by_id = {p["id"]: idx for idx, p in enumerate(catalog)}
        for er, product, _ in matches:
            if product["id"] in catalog_by_id:
                catalog[catalog_by_id[product["id"]]] = product
            else:
                catalog.append(product)
        CATALOG_PATH.write_text(json.dumps(catalog, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        print(f"catalog upserted matches; catalog size={len(catalog)}")

    if unmatched:
        print("UNMATCHED:")
        for name in unmatched:
            print(" -", name)


if __name__ == "__main__":
    main()
